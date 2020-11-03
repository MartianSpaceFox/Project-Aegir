import openpyxl as xl
from openpyxl import load_workbook
import os
import datetime

#Checks the current date and creates date object to be used and added in new workbook filename.
x = datetime.datetime.now()


#ABOUT: Script to copy multiple sheets from multiple workbooks, into a new combined workbook to make Find easier for business client.
#Case example for a cargo shipping company using Excel files.


#Add Workbooks and Worksheets

#Add the Manifest Workbook usually named MONTH YEAR in shared network folder.
fileSelect1 = input('Select Workbook 1: ')
workbook1 = xl.load_workbook(fileSelect1, data_only=True)
#Name of the sheet to be copied into the combined workbook, MANIFAST always the first
worksheet1 = workbook1['MANIFAST']

#Add the first load list file for the next trip, usually named VESSEL NAME V###.
fileSelect2 = input('Select Workbook 2: ')
workbook2 = xl.load_workbook(fileSelect2, data_only=True)
# Add sheets from file.
worksheet2_1 = workbook2['SSSM-CARS PG1']
worksheet2_2 = workbook2['SSSM-CARS PG2']
worksheet2_3 = workbook2['EXPRESS']
worksheet2_4 = workbook2['RTG1-SML']
worksheet2_5 = workbook2['RTG2-SML']
worksheet2_6 = workbook2['RTG3-SML']
worksheet2_7 = workbook2['RTG4-SML']
worksheet2_8 = workbook2['RTG5-SML']
worksheet2_9 = workbook2['RTG6-SML']
worksheet2_10 = workbook2['RTG7-SML ']
worksheet2_11 = workbook2['SSRTG1-BIG']
worksheet2_12 = workbook2['RTG1-BIG']
worksheet2_13 = workbook2['RTG2-BIG']
worksheet2_14 = workbook2['SSRTG1-CONT']
worksheet2_15 = workbook2['SSRTG2-CONT']
worksheet2_16= workbook2['RTG1-CONT']
worksheet2_17 = workbook2['RTG1-PLTS']
worksheet2_18 = workbook2['RTG2-PLTS']


#Add the second upcoming trip's load list file
fileSelect3 = input('Select Workbook 3: ')
workbook3 = xl.load_workbook(fileSelect3, data_only=True)
# Add sheets from file.
worksheet3_1 = workbook3['SSSM-CARS PG1']
worksheet3_2 = workbook3['SSSM-CARS PG2']
worksheet3_3 = workbook3['EXPRESS']
worksheet3_4 = workbook3['EXPRESS-PG 2']
worksheet3_5 = workbook3['RTG1-SML']
worksheet3_6 = workbook3['RTG2-SML']
worksheet3_7 = workbook3['RTG3-SML']
worksheet3_8 = workbook3['RTG4-SML']
worksheet3_9 = workbook3['RTG5-SML']
worksheet3_10 = workbook3['RTG6-SML']
worksheet3_11 = workbook3['RTG7-SML']
worksheet3_12 = workbook3['RTG1-BIG']
worksheet3_13 = workbook3['RTG2-BIG']
worksheet3_14 = workbook3['RTG1-CONT']
worksheet3_15 = workbook3['RTG1-PLTS']
worksheet3_16 = workbook3['RTG2-PLTS']
#worksheet3_17 = workbook3['RTG1-PLTS']
#worksheet3_18 = workbook3['RTG2-PLTS']



#New Combined Worksheet
new_workbook = xl.load_workbook('Combined4Search.xlsx')

#Creates sheet for MANIFAST copy.
wsNew1 = new_workbook.create_sheet(worksheet1.title)

#Creates Sheets for Workbook2 load list copies.
wsNew2_1= new_workbook.create_sheet(worksheet2_1.title)
wsNew2_2 = new_workbook.create_sheet(worksheet2_2.title)
wsNew2_3 = new_workbook.create_sheet(worksheet2_3.title)
wsNew2_4 = new_workbook.create_sheet(worksheet2_4.title)
wsNew2_5 = new_workbook.create_sheet(worksheet2_5.title)
wsNew2_6= new_workbook.create_sheet(worksheet2_6.title)
wsNew2_7 = new_workbook.create_sheet(worksheet2_7.title)
wsNew2_8 = new_workbook.create_sheet(worksheet2_8.title)
wsNew2_9 = new_workbook.create_sheet(worksheet2_9.title)
wsNew2_10 = new_workbook.create_sheet(worksheet2_10.title)
wsNew2_11= new_workbook.create_sheet(worksheet2_11.title)
wsNew2_12 = new_workbook.create_sheet(worksheet2_12.title)
wsNew2_13 = new_workbook.create_sheet(worksheet2_13.title)
wsNew2_14 = new_workbook.create_sheet(worksheet2_14.title)
wsNew2_15 = new_workbook.create_sheet(worksheet2_15.title)
wsNew2_16= new_workbook.create_sheet(worksheet2_16.title)
wsNew2_17 = new_workbook.create_sheet(worksheet2_17.title)
wsNew2_18 = new_workbook.create_sheet(worksheet2_18.title)

#Creates Sheets for Workbook3 load list copies.
wsNew3_1= new_workbook.create_sheet(worksheet3_1.title)
wsNew3_2 = new_workbook.create_sheet(worksheet3_2.title)
wsNew3_3 = new_workbook.create_sheet(worksheet3_3.title)
wsNew3_4 = new_workbook.create_sheet(worksheet3_4.title)
wsNew3_5 = new_workbook.create_sheet(worksheet3_5.title)
wsNew3_6= new_workbook.create_sheet(worksheet3_6.title)
wsNew3_7 = new_workbook.create_sheet(worksheet3_7.title)
wsNew3_8 = new_workbook.create_sheet(worksheet3_8.title)
wsNew3_9 = new_workbook.create_sheet(worksheet3_9.title)
wsNew3_10 = new_workbook.create_sheet(worksheet3_10.title)
wsNew3_11= new_workbook.create_sheet(worksheet3_11.title)
wsNew3_12 = new_workbook.create_sheet(worksheet3_12.title)
wsNew3_13 = new_workbook.create_sheet(worksheet3_13.title)
wsNew3_14 = new_workbook.create_sheet(worksheet3_14.title)
wsNew3_15 = new_workbook.create_sheet(worksheet3_15.title)
wsNew3_16= new_workbook.create_sheet(worksheet3_16.title)
#wsNew3_17 = new_workbook.create_sheet(worksheet3_17.title)
#wsNew3_18 = new_workbook.create_sheet(worksheet3_18.title)

#Creates Sheets for Workbook4 load list copies.
#wsNew4 = new_workbook.create_sheet(worksheet4_1.title)

#Adds the data into the first cell of the new sheet.
for row in worksheet1:
   for cell in row:
        wsNew1[cell.coordinate].value = cell.value

for row in worksheet2_1:
   for cell in row:
        wsNew2_1[cell.coordinate].value = cell.value
        wsNew2_2[cell.coordinate].value = cell.value
        wsNew2_3[cell.coordinate].value = cell.value
        wsNew2_4[cell.coordinate].value = cell.value
        wsNew2_5[cell.coordinate].value = cell.value
        wsNew2_6[cell.coordinate].value = cell.value
        wsNew2_7[cell.coordinate].value = cell.value
        wsNew2_8[cell.coordinate].value = cell.value
        wsNew2_9[cell.coordinate].value = cell.value
        wsNew2_10[cell.coordinate].value = cell.value
        wsNew2_11[cell.coordinate].value = cell.value
        wsNew2_12[cell.coordinate].value = cell.value
        wsNew2_13[cell.coordinate].value = cell.value
        wsNew2_14[cell.coordinate].value = cell.value
        wsNew2_15[cell.coordinate].value = cell.value
        wsNew2_16[cell.coordinate].value = cell.value
        wsNew2_17[cell.coordinate].value = cell.value
        wsNew2_18[cell.coordinate].value = cell.value

for row in workbook3:
    for row in worksheet3_1:
       for cell in row:
           wsNew3_1[cell.coordinate].value = cell.value
           wsNew3_2[cell.coordinate].value = cell.value
           wsNew3_3[cell.coordinate].value = cell.value
           wsNew3_4[cell.coordinate].value = cell.value
           wsNew3_5[cell.coordinate].value = cell.value
           wsNew3_6[cell.coordinate].value = cell.value
           wsNew3_7[cell.coordinate].value = cell.value
           wsNew3_8[cell.coordinate].value = cell.value
           wsNew3_9[cell.coordinate].value = cell.value
           wsNew3_10[cell.coordinate].value = cell.value
           wsNew3_11[cell.coordinate].value = cell.value
           wsNew3_12[cell.coordinate].value = cell.value
           wsNew3_13[cell.coordinate].value = cell.value
           wsNew3_14[cell.coordinate].value = cell.value
           wsNew3_15[cell.coordinate].value = cell.value
           wsNew3_16[cell.coordinate].value = cell.value
           #wsNew3_17[cell.coordinate].value = cell.value
           #wsNew3_18[cell.coordinate].value = cell.value

        #for row in workbook4:
    #for row in worksheet3:
       #for cell in row:
           #wsNew3_1[cell.coordinate].value = cell.value
           #wsNew3_2[cell.coordinate].value = cell.value
           #wsNew3_3[cell.coordinate].value = cell.value
           #wsNew3_4[cell.coordinate].value = cell.value
           #wsNew3_5[cell.coordinate].value = cell.value
           #wsNew3_6[cell.coordinate].value = cell.value
           #wsNew3_7[cell.coordinate].value = cell.value
           #wsNew3_8[cell.coordinate].value = cell.value
           #wsNew3_9[cell.coordinate].value = cell.value
           #wsNew3_10[cell.coordinate].value = cell.value
           #wsNew3_11[cell.coordinate].value = cell.value
           #wsNew3_12[cell.coordinate].value = cell.value
           #wsNew3_13[cell.coordinate].value = cell.value
           #wsNew3_14[cell.coordinate].value = cell.value
           #wsNew3_15[cell.coordinate].value = cell.value
           #wsNew3_16[cell.coordinate].value = cell.value
           #wsNew3_17[cell.coordinate].value = cell.value
           #wsNew3_18[cell.coordinate].value = cell.value

new_workbook.save ('Combined4Search.xlsx')

#Checks to see if file already exists, if it does it will copy and save a new version adding the current datetime to file name.
#if new_workbook = true
    #x.strftime("%x")
    #new_workbook = new_workbook+datetime

file = open('Combined4Search.xlsx', 'r')
#print(file.read())




